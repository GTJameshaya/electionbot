import os
import asyncio
import logging
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

# Настройка логирования
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Загрузка переменных окружения
load_dotenv()

# === Конфиги из .env ===
TOKEN = os.getenv("BOT_TOKEN")
GROUP_CHAT_ID = os.getenv("GROUP_CHAT_ID")
TARGET_BOT_TAG = os.getenv("TARGET_BOT_TAG", "@supp0rt_dom")

if not TOKEN:
    raise ValueError("BOT_TOKEN не задан в .env")
if not GROUP_CHAT_ID:
    raise ValueError("GROUP_CHAT_ID не задан в .env")
GROUP_CHAT_ID = int(GROUP_CHAT_ID)

# === Пути ===
BASE_DIR = Path(__file__).parent
TEMPLATES = {
    1: BASE_DIR / "template_1.xlsx",
    2: BASE_DIR / "template_2.xlsx"
}
ARCHIVE_DIR = BASE_DIR / "archive"
SESSIONS_DIR = BASE_DIR / "sessions"   # для временных файлов при генерации
DB_PATH = BASE_DIR / "data.db"

# Создаём директории
ARCHIVE_DIR.mkdir(exist_ok=True)
SESSIONS_DIR.mkdir(exist_ok=True)

# === Константы ===
MONTHS_RU = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
             "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
HOUSE_NAMES = {1: "Волкова д.9 к.1", 2: "Химиков 4"}
HOUSE_HASHTAG = {1: "#Волкова", 2: "#Химиков"}

BASE_KEYBOARD = ReplyKeyboardMarkup(
    [[KeyboardButton("Начать новую сессию")]],
    resize_keyboard=True
)

# ========================== РАБОТА С БАЗОЙ ДАННЫХ ==========================

def get_db_connection():
    """Возвращает соединение с SQLite и включает поддержку foreign keys."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Создаёт таблицы и заполняет справочники, если их нет."""
    conn = get_db_connection()
    cur = conn.cursor()

    # Таблица домов
    cur.execute("""
        CREATE TABLE IF NOT EXISTS houses (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            hashtag TEXT NOT NULL
        )
    """)

    # Таблица квартир
    cur.execute("""
        CREATE TABLE IF NOT EXISTS apartments (
            id INTEGER PRIMARY KEY,
            house_id INTEGER NOT NULL,
            number TEXT NOT NULL,
            sort_order INTEGER,  -- для сохранения порядка из шаблона
            FOREIGN KEY (house_id) REFERENCES houses(id),
            UNIQUE(house_id, number)
        )
    """)

    # Таблица показаний (история по месяцам)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS readings (
            id INTEGER PRIMARY KEY,
            apartment_id INTEGER NOT NULL,
            month DATE NOT NULL,   -- храним как 'YYYY-MM-DD'
            value REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (apartment_id) REFERENCES apartments(id),
            UNIQUE(apartment_id, month)
        )
    """)

    # Таблица активных сессий
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            user_id INTEGER PRIMARY KEY,
            house_id INTEGER NOT NULL,
            started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (house_id) REFERENCES houses(id)
        )
    """)

    # Заполняем дома
    cur.execute("INSERT OR IGNORE INTO houses (id, name, hashtag) VALUES (1, ?, ?)",
                (HOUSE_NAMES[1], HOUSE_HASHTAG[1]))
    cur.execute("INSERT OR IGNORE INTO houses (id, name, hashtag) VALUES (2, ?, ?)",
                (HOUSE_NAMES[2], HOUSE_HASHTAG[2]))

    conn.commit()

    # Синхронизируем квартиры из шаблонов (если они ещё не занесены)
    for house_id in (1, 2):
        sync_apartments_from_template(house_id, conn)

    conn.close()
    logger.info("База данных инициализирована")

def sync_apartments_from_template(house_id, conn):
    """Читает шаблон Excel и добавляет квартиры в таблицу apartments, если их нет."""
    template_path = TEMPLATES.get(house_id)
    if not template_path or not template_path.exists():
        logger.warning(f"Шаблон для дома {house_id} не найден")
        return

    wb = load_workbook(template_path, data_only=True)
    ws = wb.active

    cur = conn.cursor()
    # Получаем существующие номера для этого дома
    cur.execute("SELECT number FROM apartments WHERE house_id = ?", (house_id,))
    existing = {row[0] for row in cur.fetchall()}

    # Читаем номера из шаблона (столбец A, начиная с 5 строки)
    row_num = 5
    while True:
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val is None:
            break
        apt_number = str(cell_val).strip()
        if apt_number and apt_number not in existing:
            # Вставляем с sort_order = row_num
            cur.execute(
                "INSERT INTO apartments (house_id, number, sort_order) VALUES (?, ?, ?)",
                (house_id, apt_number, row_num)
            )
            existing.add(apt_number)
            logger.info(f"Добавлена квартира {apt_number} для дома {house_id}")
        row_num += 1

    conn.commit()
    wb.close()

def get_active_session(user_id):
    """Возвращает house_id, если у пользователя есть активная сессия, иначе None."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT house_id FROM sessions WHERE user_id = ?", (user_id,))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None

def create_session(user_id, house_id):
    """Создаёт новую сессию для пользователя (заменяет старую, если была)."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("REPLACE INTO sessions (user_id, house_id) VALUES (?, ?)", (user_id, house_id))
    conn.commit()
    conn.close()

def delete_session(user_id):
    """Удаляет сессию пользователя."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
    conn.commit()
    conn.close()

def save_reading(apartment_id, month, value):
    """Сохраняет показание за указанный месяц (заменяет при конфликте)."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO readings (apartment_id, month, value) VALUES (?, ?, ?) "
        "ON CONFLICT(apartment_id, month) DO UPDATE SET value = excluded.value",
        (apartment_id, month, value)
    )
    conn.commit()
    conn.close()

def get_apartments(house_id):
    """Возвращает список (id, number) для дома, отсортированный по sort_order."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, number FROM apartments WHERE house_id = ? ORDER BY sort_order",
        (house_id,)
    )
    rows = cur.fetchall()
    conn.close()
    return [(row[0], row[1]) for row in rows]

def get_reading(apartment_id, month):
    """Возвращает значение показания для квартиры за указанный месяц, или None."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT value FROM readings WHERE apartment_id = ? AND month = ?",
                (apartment_id, month))
    row = cur.fetchone()
    conn.close()
    return row[0] if row else None

# ========================== ГЕНЕРАЦИЯ EXCEL ==========================

def generate_excel(house_id, user_id):
    """
    Создаёт Excel-файл на основе шаблона и данных из БД.
    Возвращает путь к временному файлу.
    """
    template_path = TEMPLATES.get(house_id)
    if not template_path or not template_path.exists():
        raise FileNotFoundError(f"Шаблон для дома {house_id} не найден")

    wb = load_workbook(template_path)
    ws = wb.active

    today = datetime.now()
    current_month = today.replace(day=1).date()
    prev_month = (current_month - timedelta(days=1)).replace(day=1)

    # Заполняем заголовки
    ws.cell(row=4, column=4, value=MONTHS_RU[prev_month.month - 1])
    ws.cell(row=4, column=5, value=MONTHS_RU[current_month.month - 1])
    ws["D2"] = current_month.replace(day=21)   # дата 21-го числа текущего месяца

    # Получаем квартиры
    apartments = get_apartments(house_id)

    row = 5
    for apt_id, apt_num in apartments:
        # Предыдущее показание
        prev_val = get_reading(apt_id, prev_month)
        # Текущее показание
        cur_val = get_reading(apt_id, current_month)

        ws.cell(row=row, column=1, value=apt_num)
        if prev_val is not None:
            ws.cell(row=row, column=4, value=prev_val)
        if cur_val is not None:
            ws.cell(row=row, column=5, value=cur_val)
            if prev_val is not None:
                ws.cell(row=row, column=6, value=cur_val - prev_val)
            else:
                ws.cell(row=row, column=6, value=0)  # или оставить пустым
        else:
            # Если текущего нет, можно оставить пустым
            pass

        row += 1

    # Сохраняем во временный файл
    temp_file = SESSIONS_DIR / f"temp_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(temp_file)
    wb.close()
    return temp_file

# ========================== ОБРАБОТЧИКИ КОМАНД ==========================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда /start – предложить выбрать дом."""
    keyboard = [[KeyboardButton(HOUSE_NAMES[1]), KeyboardButton(HOUSE_NAMES[2])]]
    await update.message.reply_text(
        "Выбери дом:",
        reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка всех текстовых сообщений."""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    logger.info(f"Получено сообщение от {user_id}: {text}")

    # 1. Кнопка "Начать новую сессию"
    if text == "Начать новую сессию":
        await start(update, context)
        return

    # 2. Выбор дома (если текст совпадает с названием)
    if text in HOUSE_NAMES.values():
        house_id = 1 if text == HOUSE_NAMES[1] else 2
        # Проверяем наличие шаблона
        if not TEMPLATES[house_id].exists():
            await update.message.reply_text("Нет шаблона для этого дома")
            return

        # Создаём сессию
        create_session(user_id, house_id)
        keyboard = [[KeyboardButton("Файл сейчас"), KeyboardButton("Готово")]]
        await update.message.reply_text(
            f"Дом: {text}\nСессия начата — вводи показания\nТишина = всё ок",
            reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        )
        return

    # 3. "Файл сейчас" – выгрузить текущее состояние
    if text == "Файл сейчас":
        house_id = get_active_session(user_id)
        if house_id is None:
            await update.message.reply_text("Нет активной сессии. Начните новую.")
            return
        try:
            temp_file = generate_excel(house_id, user_id)
            await update.message.reply_document(
                open(temp_file, "rb"),
                caption="Текущий файл (показания на данный момент)"
            )
            # Удаляем временный файл после отправки (опционально)
            temp_file.unlink(missing_ok=True)
        except Exception as e:
            logger.error(f"Ошибка генерации файла: {e}")
            await update.message.reply_text(f"❌ Ошибка: {str(e)}")
        return

    # 4. "Готово" – завершить сессию и отправить финальный файл в группу
    if text == "Готово":
        house_id = get_active_session(user_id)
        if house_id is None:
            await update.message.reply_text("Нет активной сессии.")
            return

        try:
            # Генерируем финальный Excel
            temp_file = generate_excel(house_id, user_id)

            # Перемещаем в архив с постоянным именем
            now = datetime.now()
            month_tag = f"{MONTHS_RU[now.month - 1]}_{now.year}"
            archive_name = ARCHIVE_DIR / f"{now.strftime('%Y%m%d_%H%M')}_house{house_id}_{HOUSE_NAMES[house_id].replace(' ', '_')}.xlsx"
            temp_file.rename(archive_name)

            # Отправляем в группу
            with open(archive_name, "rb") as f:
                caption = (
                    f"{TARGET_BOT_TAG} Привет! Показания за текущий {now.strftime('%m.%Y')} — {HOUSE_NAMES[house_id]}\n"
                    f"#{month_tag} {HOUSE_HASHTAG[house_id]}"
                )
                await context.bot.send_document(
                    GROUP_CHAT_ID,
                    f,
                    filename=archive_name.name,
                    caption=caption
                )

            await update.message.reply_text(
                "✅ Файл улетел в группу. Сессия завершена.",
                reply_markup=BASE_KEYBOARD
            )

            # Удаляем сессию
            delete_session(user_id)

        except Exception as e:
            logger.error(f"Ошибка при завершении сессии: {e}")
            await update.message.reply_text(f"❌ Ошибка: {str(e)}")
        return

    # 5. Ввод показаний (если есть активная сессия)
    house_id = get_active_session(user_id)
    if house_id is None:
        await update.message.reply_text(
            "Сначала начни сессию: /start",
            reply_markup=BASE_KEYBOARD
        )
        return

    # Парсим пары "квартира показание"
    parts = text.replace("\n", " ").split()
    if len(parts) % 2 != 0:
        await update.message.reply_text("❌ Нечётное количество значений. Введите как: 1 100 2 200")
        return

    # Получаем список квартир для этого дома для проверки
    apartments = get_apartments(house_id)
    apt_map = {num: apt_id for apt_id, num in apartments}

    current_month = datetime.now().replace(day=1).date()
    updated_count = 0
    errors = []

    for i in range(0, len(parts), 2):
        apt_num = parts[i].strip()
        try:
            val = float(parts[i+1])  # разрешаем дробные
        except ValueError:
            errors.append(f"Для квартиры {apt_num} введите число")
            continue

        apt_id = apt_map.get(apt_num)
        if apt_id is None:
            errors.append(f"Квартира {apt_num} не найдена в этом доме")
        else:
            save_reading(apt_id, current_month, val)
            updated_count += 1

    # Отправляем ответ
    if updated_count > 0:
        await update.message.reply_text(f"✅ Обновлено {updated_count} квартир")
    if errors:
        await update.message.reply_text("❌ " + "\n".join(errors[:5]))  # ограничим вывод

# ========================== ЗАПУСК БОТА ==========================

def main():
    """Инициализация БД и запуск бота."""
    init_db()

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("🚀 Бот запущен!")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
